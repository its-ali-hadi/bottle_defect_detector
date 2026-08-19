from __future__ import annotations

import argparse
import json
import shutil
from collections import Counter
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image as PILImage

from bottle_detector.models import LABELS_AR


DEFAULT_INPUT = Path("outputs/detections.json")
DEFAULT_OUTPUT_DIR = Path("result")
LATEST_OUTPUT_NAME = "detections_latest.xlsx"

BOTTLE_DETAIL_HEADERS = [
    "تسلسل العلبة",
    "عدد العيوب",
    "العيوب بالتفصيل",
    "ملخص العلبة",
    "Accuracy %",
    "Precision %",
    "Recall %",
    "صورة العلبة",
]
BOTTLE_IMAGE_COLUMN = len(BOTTLE_DETAIL_HEADERS)
THUMBNAIL_WIDTH_PX = 160
ROW_HEIGHT_PADDING_PX = 12
PX_TO_POINTS = 0.75
IMAGE_COLUMN_WIDTH = 24


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert bottle detector JSON results to an Excel workbook.",
    )
    parser.add_argument(
        "--input",
        default=str(DEFAULT_INPUT),
        help="Input detections JSON path. Default: outputs/detections.json",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Output directory. Default: result",
    )
    parser.add_argument(
        "--output-name",
        default=None,
        help="Optional exact Excel file name. If omitted, creates a timestamped file and detections_latest.xlsx.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    result = export_excel_report(
        input_path=Path(args.input),
        output_dir=Path(args.output_dir),
        output_name=args.output_name,
    )
    print(f"Wrote Excel report: {result['report_path']}")
    if result.get("latest_path"):
        print(f"Updated latest Excel report: {result['latest_path']}")
    return 0


def export_excel_report(
    *,
    input_path: Path = DEFAULT_INPUT,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    output_name: str | None = None,
) -> dict[str, Path | None]:
    data = load_json(input_path)
    workbook = build_workbook(data, input_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    report_path = output_dir / (output_name or timestamped_output_name())
    workbook.save(report_path)

    latest_path: Path | None = None
    if output_name is None:
        latest_path = output_dir / LATEST_OUTPUT_NAME
        try:
            shutil.copyfile(report_path, latest_path)
        except PermissionError:
            latest_path = None

    return {"report_path": report_path, "latest_path": latest_path}


def timestamped_output_name() -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"detections_{stamp}.xlsx"


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Input JSON not found: {path}")
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("Input JSON must contain an object at the top level.")
    if not isinstance(payload.get("detections"), list):
        raise ValueError('Input JSON must contain a "detections" list.')
    return payload


def build_workbook(data: dict[str, Any], input_path: Path) -> Workbook:
    workbook = Workbook()
    report_sheet = workbook.active
    report_sheet.title = "Report"

    detections = data.get("detections", [])
    add_full_report_sheet(report_sheet, data, detections, input_path)
    summary_sheet = workbook.create_sheet("Summary")
    add_summary_sheet(summary_sheet, data, detections)
    add_bottle_details_sheet(workbook, detections, input_path)

    image_columns = {"Report": BOTTLE_IMAGE_COLUMN, "Bottle Details": BOTTLE_IMAGE_COLUMN}
    for sheet in workbook.worksheets:
        sheet.sheet_view.rightToLeft = True
        apply_common_styles(sheet, image_column=image_columns.get(sheet.title))
    return workbook


def add_full_report_sheet(
    sheet: Any,
    data: dict[str, Any],
    detections: list[dict[str, Any]],
    input_path: Path,
) -> None:
    defect_counts = Counter(
        defect.get("type", "unknown")
        for item in detections
        for defect in item.get("defects", [])
    )
    stats = data.get("statistics", {}) or {}
    summary_rows = [
        ("المصدر", data.get("source", "")),
        ("المودل", data.get("model", "")),
        ("تاريخ إنشاء JSON", data.get("created_at", "")),
        ("تاريخ تصدير Excel", datetime.now().isoformat(timespec="seconds")),
        ("عدد العلب الكلي", len(detections)),
        ("عيوب جسم العلبة", defect_counts.get("body_defect", 0)),
        ("علب متسخة", defect_counts.get("dirty", 0)),
        ("عيوب تصنيعية", defect_counts.get("factory_defect", 0)),
        ("عدد العلب المعيبة", stats.get("defective_count", 0)),
        ("عدد العلب السليمة", stats.get("ok_count", 0)),
        ("نسبة الدقة العامة Accuracy %", format_pct(stats.get("accuracy_pct"))),
        ("دقة العيوب المكتشفة Precision %", format_pct(stats.get("precision_pct"))),
        ("نسبة الاسترجاع Recall %", format_pct(stats.get("recall_pct"))),
    ]

    sheet.append(["الإحصائية", "القيمة"])
    for row in summary_rows:
        sheet.append(list(row))
    add_table(sheet, "ReportSummaryTable", start_row=1, end_row=sheet.max_row, end_column=2)

    sheet.append([])
    sheet.append(BOTTLE_DETAIL_HEADERS)
    details_header_row = sheet.max_row
    write_bottle_rows(sheet, detections, input_path)
    add_table(
        sheet,
        "ReportBottleDetailsTable",
        start_row=details_header_row,
        end_row=sheet.max_row,
        end_column=len(BOTTLE_DETAIL_HEADERS),
    )


def add_summary_sheet(sheet: Any, data: dict[str, Any], detections: list[dict[str, Any]]) -> None:
    defect_counts = Counter(
        defect.get("type", "unknown")
        for item in detections
        for defect in item.get("defects", [])
    )
    stats = data.get("statistics", {}) or {}
    rows = [
        ("Source", data.get("source", "")),
        ("Model", data.get("model", "")),
        ("Total Bottles", len(detections)),
        ("Body Defects", defect_counts.get("body_defect", 0)),
        ("Dirty Bottles", defect_counts.get("dirty", 0)),
        ("Factory Defects", defect_counts.get("factory_defect", 0)),
        ("Defective Bottles", stats.get("defective_count", 0)),
        ("OK Bottles", stats.get("ok_count", 0)),
        ("Accuracy %", format_pct(stats.get("accuracy_pct"))),
        ("Precision %", format_pct(stats.get("precision_pct"))),
        ("Recall %", format_pct(stats.get("recall_pct"))),
    ]
    sheet.append(["Field", "Value"])
    for row in rows:
        sheet.append(list(row))
    add_table(sheet, "SummaryTable")


def add_bottle_details_sheet(workbook: Workbook, detections: list[dict[str, Any]], input_path: Path) -> None:
    sheet = workbook.create_sheet("Bottle Details")
    sheet.append(BOTTLE_DETAIL_HEADERS)
    write_bottle_rows(sheet, detections, input_path)
    add_table(sheet, "BottleDetailsTable", end_column=len(BOTTLE_DETAIL_HEADERS))


def write_bottle_rows(sheet: Any, detections: list[dict[str, Any]], input_path: Path) -> None:
    image_column_letter = get_column_letter(BOTTLE_IMAGE_COLUMN)
    for item in detections:
        sheet.append(bottle_row(item))
        row_number = sheet.max_row
        crop_path = resolve_crop_path(item.get("crop_path", ""), input_path)
        thumbnail = build_thumbnail(crop_path) if crop_path else None
        if thumbnail is None:
            sheet.cell(row=row_number, column=BOTTLE_IMAGE_COLUMN, value="لا توجد صورة")
            continue
        xl_image, thumb_height = thumbnail
        sheet.add_image(xl_image, f"{image_column_letter}{row_number}")
        sheet.row_dimensions[row_number].height = max(
            sheet.row_dimensions[row_number].height or 0,
            (thumb_height + ROW_HEIGHT_PADDING_PX) * PX_TO_POINTS,
        )


def bottle_row(item: dict[str, Any]) -> list[Any]:
    defects = item.get("defects", [])
    return [
        item.get("sequence", ""),
        len(defects),
        format_defects_for_bottle(defects),
        item.get("summary_ar", ""),
        format_pct(item.get("accuracy_pct")),
        format_pct(item.get("precision_pct")),
        format_pct(item.get("recall_pct")),
    ]


def format_pct(value: Any) -> str:
    if value is None:
        return "-"
    return f"{value}%"


def resolve_crop_path(crop_path_str: str, input_path: Path) -> Path | None:
    if not crop_path_str:
        return None
    candidate = Path(crop_path_str)
    if candidate.is_absolute():
        return candidate if candidate.exists() else None
    project_root = input_path.resolve().parent.parent
    for base in (Path.cwd(), project_root):
        resolved = (base / candidate).resolve()
        if resolved.exists():
            return resolved
    return None


def build_thumbnail(path: Path) -> tuple[XLImage, int] | None:
    try:
        with PILImage.open(path) as source_image:
            rgb_image = source_image.convert("RGB")
            ratio = THUMBNAIL_WIDTH_PX / rgb_image.width
            thumb_height = max(1, round(rgb_image.height * ratio))
            thumbnail = rgb_image.resize((THUMBNAIL_WIDTH_PX, thumb_height))
            buffer = BytesIO()
            thumbnail.save(buffer, format="JPEG", quality=85)
            buffer.seek(0)
    except Exception:
        return None
    xl_image = XLImage(buffer)
    xl_image.width = THUMBNAIL_WIDTH_PX
    xl_image.height = thumb_height
    return xl_image, thumb_height


def add_table(
    sheet: Any,
    name: str,
    *,
    start_row: int = 1,
    end_row: int | None = None,
    end_column: int | None = None,
) -> None:
    end_row = end_row or sheet.max_row
    end_column = end_column or sheet.max_column
    end_column_letter = get_column_letter(end_column)
    ref = f"A{start_row}:{end_column_letter}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def apply_common_styles(sheet: Any, *, image_column: int | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_rows = find_header_rows(sheet)
    for row_number in header_rows:
        for cell in sheet[row_number]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2" if sheet.title != "Report" else "A11"
    auto_size_columns(sheet, skip_columns={image_column} if image_column else set())
    if image_column:
        sheet.column_dimensions[get_column_letter(image_column)].width = IMAGE_COLUMN_WIDTH


def find_header_rows(sheet: Any) -> set[int]:
    header_markers = {
        "Field",
        "Sequence",
        "الإحصائية",
        "تسلسل العلبة",
    }
    rows = {1}
    for row in sheet.iter_rows():
        first_value = row[0].value
        if first_value in header_markers:
            rows.add(row[0].row)
    return rows


def auto_size_columns(sheet: Any, skip_columns: set[int] | None = None) -> None:
    skip_columns = skip_columns or set()
    for column_cells in sheet.columns:
        column_index = column_cells[0].column
        if column_index in skip_columns:
            continue
        column_letter = get_column_letter(column_index)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 60)


def format_defects_for_bottle(defects: list[dict[str, Any]]) -> str:
    if not defects:
        return "لا توجد عيوب"
    lines: list[str] = []
    for index, defect in enumerate(defects, start=1):
        label = defect.get("label_ar") or defect_type_ar(defect.get("type", ""))
        defect_type = defect.get("type", "")
        description = defect.get("description_ar", "")
        lines.append(
            f"{index}. العيب: {label}"
            f"\n   النوع: {defect_type}"
            f"\n   الوصف: {description}"
        )
    return "\n".join(lines)


def defect_type_ar(defect_type: Any) -> str:
    return LABELS_AR.get(str(defect_type), str(defect_type))


if __name__ == "__main__":
    raise SystemExit(main())
